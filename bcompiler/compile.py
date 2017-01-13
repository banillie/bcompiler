"""
Initial Docstring.
"""
import fnmatch
import logging
import os
import re
from datetime import date

from bcompiler.datamap import Datamap

from bcompiler.process import Cleanser
from bcompiler.process.simple_comparitor import FileComparitor, ParsedMaster

from bcompiler.utils import DATAMAP_RETURN_TO_MASTER, OUTPUT_DIR, RETURNS_DIR
from bcompiler.utils import cell_bg_colour, bc_is_close, quick_typechecker

from openpyxl import load_workbook, Workbook

CELL_REGEX = re.compile('[A-Z]+[0-9]+')
DROPDOWN_REGEX = re.compile('^\D*$')
TODAY = date.today().isoformat()

logger = logging.getLogger('bcompiler.compiler')

DATA_MAP_FILE = DATAMAP_RETURN_TO_MASTER


def get_current_quarter(source_file):
    """
    DOCSTRING HERE
    """
    wb = load_workbook(RETURNS_DIR + source_file, read_only=True)
    ws = wb['Summary']
    q = ws['G3'].value
    logger.info('Getting current Quarter as {}'.format(q))
    return q


def parse_source_cells(source_file, datamap_source_file):
    """
    Doc string in here.
    """
    ls_of_dataline_dicts = []
    wb = load_workbook(source_file, read_only=True, data_only=True)
    datamap_obj = Datamap(
        datamap_type='returns-to-master',
        source_file=datamap_source_file)
    for item in datamap_obj.data:
        # hack for importation (we have a new sheet!)
        if item.sheet is not None and item.cellref is not None:
            ws = wb[item.sheet.rstrip()]
            try:
                v = ws[item.cellref.rstrip()].value
            except IndexError:
                logger.error(
                    "Datamap wants sheet: {}; cellref: {} but this is out"
                    "of range.\n\tFile: {}".format(
                        item.sheet,
                        item.cellref,
                        source_file))
                v = ""
            else:
                if v is None:
                    logger.debug(
                        "{} in {} is empty.".format(
                            item.cellref,
                            item.sheet))
                elif type(v) == str:
                    v = v.rstrip()
                else:
                    logger.debug(
                        "{} in {} is {}".format(
                            item.cellref,
                            item.sheet,
                            v))
                try:
                    c = Cleanser(v)
                except IndexError:
                    logger.error(
                        ("Trying to clean an empty cell {} at sheet {} in {}. "
                         "Ignoring.").format(
                            item.cellref, item.sheet, source_file))
                except TypeError:
                    pass
                else:
                    v = c.clean()
            destination_kv = dict(gmpp_key=item.cellname, gmpp_key_value=v)
            ls_of_dataline_dicts.append(destination_kv)
    return ls_of_dataline_dicts


def write_excel(source_file, count, workbook, compare_master=None):
    """
    count is used to count number of times function is run so that multiple
    returns can be added
    and not overwrite the GMPP key column
    let's create an Excel file in memory
    it will have one worksheet - let's get it
    """
    ws = workbook.active

    # give it a title
    ws.title = "Constructed BICC Data Master"

    # this is the data from the source spreadsheet
    out_map = parse_source_cells(source_file, DATAMAP_RETURN_TO_MASTER)

    # we need to the project name to work out index order for comparing
    # master file
    project_name = [
        item['gmpp_key_value']
        for item in out_map if item['gmpp_key'] == 'Project/Programme Name'][0]

    try:
        if compare_master:
            compare_file = compare_master[0]
    except TypeError:
        compare_file = None

    if compare_master:
        parsed_master = ParsedMaster(compare_file)
        hd_indices = parsed_master._project_header_index
        this_index = [v for k, v in hd_indices.items() if k == project_name][0]

    try:
        # this deals with issue of not passing --compare to compile argument
        comparitor = FileComparitor([compare_file])
    except UnboundLocalError:
        pass

    if count == 1:
        i = 1
        # this one writes the first column, the keys
        for d in out_map:
            c = ws.cell(row=i, column=1)
            c.value = d['gmpp_key']
            i += 1
        i = 1

        # then it writes the second column with the values
        for d in out_map:

            c = ws.cell(row=i, column=2)

            # HERE WE NEED TO KNOW TWO THINGS:
            # - name of project we're compiling
            # - index of the header for that project in the comparing
            # master, if we are comparing

            try:
                # exception will be if we call without compare flag
                compare_val = comparitor.compare(this_index, d['gmpp_key'])
            except UnboundLocalError:
                compare_val = False

            # if there is something to compare it
            if compare_val:

                # if compare_val is a valid type (float, int or date)
                # but this can change - we need to add str
                if quick_typechecker(d['gmpp_key_value'], compare_val):

                    # if there is enough of a difference in values
                    if bc_is_close(d['gmpp_key_value'], compare_val) is False:

                        # if current value is HIGHER than earlier value
                        if compare_val < d['gmpp_key_value']:

                            # ... round it
                            try:
                                # exception expected a date object
                                # which round will not handle
                                compare_val = round(compare_val, 2)
                            except TypeError:
                                pass

                            if isinstance(compare_val, (int, float)):
                                # ... fill the background cell with RED
                                c.fill = cell_bg_colour(rgb=[255, 0, 0])
                            elif isinstance(compare_val, date):
                                # ... fill the background cell with PURPLE
                                c.fill = cell_bg_colour(rgb=[255, 0, 127])

                            # ... write to the cell
                            c.value = d['gmpp_key_value']
                            if isinstance(d['gmpp_key_value'], date):
                                c.number_format = 'd/mm/yy'

                        # if current value is LOWER than earlier value
                        elif compare_val > d['gmpp_key_value']:

                            # ... round it
                            try:
                                compare_val = round(compare_val, 2)
                            except TypeError:
                                pass

                            if isinstance(compare_val, (int, float)):
                                # ... fill the background cell with GREEN
                                c.fill = cell_bg_colour(rgb=[3, 180, 0])
                            elif isinstance(compare_val, date):
                                # ... fill the background cell with GREEN
                                c.fill = cell_bg_colour(rgb=[93, 81, 0])
                                c.number_format = 'd/mm/yy'

                            # ... write to the cell
                            c.value = d['gmpp_key_value']
                            if isinstance(d['gmpp_key_value'], date):
                                c.number_format = 'd/mm/yy'

                        else:
                            # ... write to the cell
                            c.value = d['gmpp_key_value']
                            if isinstance(d['gmpp_key_value'], date):
                                c.number_format = 'd/mm/yy'

                    else:
                        # if there is no discernable difference in cells
                        # write to the cell
                        c.value = d['gmpp_key_value']
                        if isinstance(d['gmpp_key_value'], date):
                            c.number_format = 'd/mm/yy'
                else:
                    # not a type of interest for comparison
                    # write to the cell
                    c.value = d['gmpp_key_value']
                    if isinstance(d['gmpp_key_value'], date):
                        c.number_format = 'd/mm/yy'
            else:
                # if there is no compare value, just write to the cell
                # this writes to the cell
                c.value = d['gmpp_key_value']
                if isinstance(d['gmpp_key_value'], date):
                    c.number_format = 'd/mm/yy'

            i += 1
    else:
        i = 1
        # now we have no need of the keys any more so we're just writing
        # values here
        for d in out_map:
            c = ws.cell(row=i, column=count + 1)

            try:
                compare_val = comparitor.compare(this_index, d['gmpp_key'])
            except UnboundLocalError:
                compare_val = False

            # if there is something to compare it
            if compare_val:

                # if compare_val is a float or int
                if quick_typechecker(d['gmpp_key_value'], compare_val):

                    # if there is enough of a difference in values
                    if bc_is_close(d['gmpp_key_value'], compare_val) is False:

                        # if current value is HIGHER than earlier value
                        if compare_val < d['gmpp_key_value']:

                            # ... round it
                            # exception expected a date object
                            # which round will not handle
                            try:
                                compare_val = round(compare_val, 2)
                            except TypeError:
                                pass

                            if isinstance(compare_val, (int, float)):
                                # ... fill the background cell with RED
                                c.fill = cell_bg_colour(rgb=[255, 0, 0])
                            elif isinstance(compare_val, date):
                                # ... fill the background cell with PURPLE
                                c.fill = cell_bg_colour(rgb=[255, 0, 127])
                                c.number_format = 'd/mm/yy'

                            # ... write to the cell
                            c.value = d['gmpp_key_value']
                            if isinstance(d['gmpp_key_value'], date):
                                c.number_format = 'd/mm/yy'

                        # if current value is LOWER than earlier value
                        elif compare_val > d['gmpp_key_value']:

                            # ... round it
                            # exception expected a date object
                            # which round will not handle
                            try:
                                compare_val = round(compare_val, 2)
                            except TypeError:
                                pass

                            if isinstance(compare_val, (int, float)):
                                # ... fill the background cell with GREEN
                                c.fill = cell_bg_colour(rgb=[3, 180, 0])
                            elif isinstance(compare_val, date):
                                # ... fill the background cell with GREEN
                                c.fill = cell_bg_colour(rgb=[93, 81, 0])
                                c.number_format = 'd/mm/yy'

                            # ... write to the cell
                            c.value = d['gmpp_key_value']
                            if isinstance(d['gmpp_key_value'], date):
                                c.number_format = 'd/mm/yy'

                        else:
                            # ... write to the cell
                            c.value = d['gmpp_key_value']
                            if isinstance(d['gmpp_key_value'], date):
                                c.number_format = 'd/mm/yy'

                    else:
                        # if there is no discernable difference in cells
                        # write to the cell
                        c.value = d['gmpp_key_value']
                        if isinstance(d['gmpp_key_value'], date):
                            c.number_format = 'd/mm/yy'
                else:
                    # if there is no discernable difference in cells
                    # write to the cell
                    c.value = d['gmpp_key_value']
                    if isinstance(d['gmpp_key_value'], date):
                        c.number_format = 'd/mm/yy'
            else:
                # if there is no compare value, just write to the cell
                # this writes to the cell
                c.value = d['gmpp_key_value']
                if isinstance(d['gmpp_key_value'], date):
                    c.number_format = 'd/mm/yy'

            i += 1


def run(compare_master=None):
    """
    Run the compile function.
    """
    # if we want to do a comparison
    if compare_master:

        workbook = Workbook()
        count = 1
        for file in os.listdir(RETURNS_DIR):
            if fnmatch.fnmatch(file, '*.xlsx'):
                logger.info("Processing {}".format(file))
                write_excel(
                    (RETURNS_DIR + file),
                    count=count,
                    workbook=workbook,
                    compare_master=compare_master
                )
                count += 1
        for file in os.listdir(RETURNS_DIR):
            cq = get_current_quarter(file)
            if cq is not None:
                break
        OUTPUT_FILE = '{}compiled_master_{}_{}.xlsx'.format(
            OUTPUT_DIR, TODAY, cq)
        workbook.save(OUTPUT_FILE)
    else:
        # we just want a straight master with no change indication
        workbook = Workbook()
        count = 1
        for file in os.listdir(RETURNS_DIR):
            if fnmatch.fnmatch(file, '*.xlsx'):
                logger.info("Processing {}".format(file))
                write_excel(
                    (RETURNS_DIR + file),
                    count=count,
                    workbook=workbook,
                )
                count += 1
        for file in os.listdir(RETURNS_DIR):
            cq = get_current_quarter(file)
            if cq is not None:
                break
        OUTPUT_FILE = '{}compiled_master_{}_{}.xlsx'.format(
            OUTPUT_DIR, TODAY, cq)
        workbook.save(OUTPUT_FILE)
