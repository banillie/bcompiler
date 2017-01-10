import pytest

from bcompiler.process.simple_comparitor import BCCell
from bcompiler.process.simple_comparitor import populate_cells

from openpyxl import Workbook


key_col_data = [
    'Project/Programme Name',
    'SRO Sign-Off',
    'Reporting period (GMPP - Snapshot Date)',
    'Quarter Joined',
    'GMPP (GMPP – formally joined GMPP)',
    'IUK top 40',
    'Top 37',
    'DfT Business Plan',
    'GMPP - IPA ID Number',
    'DFT ID Number',
    'Working Contact Name',
    'Working Contact Telephone',
    'Working Contact Email',
    'DfT Group',
    'DfT Division',
    'Agency or delivery partner (GMPP - Delivery Organisation primary)',
    'Strategic Alignment/Government Policy (GMPP – Key drivers)',
    'Project Scope',
    'Brief project description (GMPP – brief descripton)',
    'Delivery Structure',
    'Description if \'Other',
    'Change Delivery Methodology'
]

project_b_data = [
    'Digital Signalling',
    '2016-10-12 0:00:00',
    'Q2 1617',
    None,
    None,
    None,
    None,
    None,
    None,
    8,
    'Niall Le Mage',
    '2079442043',
    'niall.lemage@dft.gsi.gov.uk',
    'Rail Group',
    'Network Services',
    'Network Rail',
    'In line with DfTs single Departmental Plan to roll out new technology',
    'Scope of the ETCS cab-fitment fund: | to facilitate the inclusion of',
    'The fitting of digital signalling technology to prototype passenger',
    'Project',
    None,
    'Waterfall',
]


@pytest.fixture
def populate_test_data():
    wb = Workbook()
    ws = wb.active

    for item in key_col_data:
        c = BCCell(item, row_num=key_col_data.index(item) + 1, col_num=1)
        ws.cell(value=c.value, row=c.row_num, column=c.col_num)
    yield ws


def test_populate_test_data(populate_test_data):
    assert populate_test_data['A1'].value == 'Project/Programme Name'
    assert populate_test_data['A2'].value == 'SRO Sign-Off'


def test_populate_function():
    wb = Workbook()
    ws = wb.active

    # populate by coordinates
    populate_cells(ws, [BCCell("Test1", cellref="A1")])
    assert ws['A1'].value == "Test1"

    # populate by row, col
    populate_cells(ws, [BCCell("Fanciso Monk", 2, 1)])
    assert ws['A2'].value == "Fanciso Monk"
