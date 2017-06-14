import os
import pytest
import unittest

import re

from bcompiler.compile import parse_source_cells, get_current_quarter
from bcompiler.datamap import Datamap, DatamapLine

from bcompiler.utils import VALIDATION_REFERENCES, SHEETS
from bcompiler.utils import index_returns_directory

from openpyxl import Workbook


def test_parse_source_excel_file(bicc_return, mock_datamap_source_file):
    import pudb; pudb.set_trace()  # XXX BREAKPOINT
    parsed_data = parse_source_cells(bicc_return, mock_datamap_source_file)
    assert 'Project/Programme Name' == parsed_data[0]['gmpp_key']
    assert 'DVSA IT Sourcing' == parsed_data[0]['gmpp_key_value']


def test_dropdown_not_passing_to_master_bug(bicc_return, mock_datamap_source_file):
    data = parse_source_cells(bicc_return, mock_datamap_source_file)
    example_validated_cell = "IO1 - Monetised?"
    matches = [x for x in data if x['gmpp_key'] == example_validated_cell]
    assert matches[0]['gmpp_key'] == example_validated_cell


def test_parse_returned_form(bicc_return, mock_datamap_source_file):
    data = parse_source_cells(bicc_return, mock_datamap_source_file)
    assert data[0]['gmpp_key'] == 'Project/Programme Name'


class TestDatamapFunctionality(unittest.TestCase):
    def setUp(self):
        self.cell_regex = re.compile('[A-Z]+[0-9]+')
        self.dropdown_regex = re.compile('^\D*$')
        self.docs = os.path.join(os.path.expanduser('~'), 'Documents')
        bcomp_working_d = 'bcompiler'
        self.path = os.path.join(self.docs, bcomp_working_d)
        self.source_path = os.path.join(self.path, 'source')
        self.output_path = os.path.join(self.path, 'output')
        self.datamap_master_to_returns = os.path.join(
            self.source_path, 'datamap-master-to-returns')
        self.datamap_returns_to_master = os.path.join(
            self.source_path, 'datamap-returns-to-master')
        self.master = os.path.join(self.source_path, 'master.csv')
        self.transposed_master = os.path.join(
            self.source_path, 'master_transposed.csv')
        self.dm = Datamap(
            datamap_type='returns-to-master',
            source_file=self.datamap_returns_to_master)

    def test_verified_lines(self):
        # these are DatamapLine objects that have 4 attributes, the last of
        # which is verification dropdown text
        # the last element in each should be a dropdown text string
        for item in self.dm._dml_cname_sheet_cref_ddown:
            self.assertTrue(self.dropdown_regex, item.dropdown_txt)

    def test_verified_lines_for_dropdown_text(self):
        # we're expecting the dropdown_txt attr in the DatamapLine object
        # to be what we expect
        for item in self.dm._dml_cname_sheet_cref_ddown:
            self.assertTrue(item.dropdown_txt in VALIDATION_REFERENCES.keys())

    def test_non_verified_lines(self):
        # these are DatamapLine objects that have 3 attributes, the
        # last of which is a regex
        for item in self.dm._dml_cname_sheet_cref:
            self.assertTrue(self.cell_regex, item.cellref)

    def test_cells_that_will_not_migrate(self):
        # these are DatamapLine objects that have 2 attributes,
        # the last of which is a sheet ref
        for item in self.dm._dml_cname_sheet:
            self.assertTrue(item.sheet in SHEETS)

    def test_single_item_lines(self):
        # DatamapLines that have a single attribute
        for item in self.dm._dml_cname:
            # TODO this is fragile - shouldn't be counting lines in this test
            self.assertEqual(self.dm.count_dml_cellname_only, 18)

    def test_datamap_is_cleaned_attr(self):
        self.assertTrue(self.dm.is_cleaned)

    def test_pretty_dataline_print(self):
        dml = DatamapLine()
        dml.cellname = 'Test cellname'
        dml.sheet = 'Summary'
        dml.cellref = 'C12'
        dml.dropdown_txt = 'Finance Figures'
        self.assertEqual(
            dml.pretty_print(), ("Name: Test cellname | Sheet: Summary | "
                                 "Cellref: C12 | Dropdown: Finance Figures"))

    def test_index_returns_directory(self):
        assert index_returns_directory() == []


# LATER PYTEST-ONLY TESTS FOR COMPILATION PROCESS, USING FIXTURES

@pytest.fixture
def bicc_return():
    wb = Workbook()
    wb.create_sheet('Summary')
    wb.create_sheet('Approval & Project milestones')
    wb.create_sheet('Finance & Benefits')
    wb.create_sheet('Resources')
    wb.create_sheet('Assurance planning')
    wb.create_sheet('GMPP info')
    ws_summary = wb['Summary']
    # enter some values in the right slots
    ws_summary['B5'].value = 'Cookfield Rebuild'
    ws_summary['B8'].value = 'Roads, Monitoring and Horse'
    ws_summary['C12'].value = 'Keith McGilvrey'
    ws_summary['C13'].value = '0208 944 3554'
    ws_summary['C14'].value = 'cleft.palet@screen.co.ok'
    ws_summary['C21'].value = 'Cohort 13'
    ws_summary['B26'].value = 'We think this is a really good idea, ok?'

    wb.save('/tmp/test-bicc-return.xlsx')
    yield '/tmp/test-bicc-return.xlsx'
    os.unlink('/tmp/test-bicc-return.xlsx')


@pytest.fixture
def old_master():
    wb = Workbook()
    wb.create_sheet('Constructed BICC Data Master')
    ws = wb['Constructed BICC Data Master']
    ws['A1'].value = 'Project/Programme Name'
    ws['A2'].value = 'SRO Sign-Off'
    # TODO carry on with this
