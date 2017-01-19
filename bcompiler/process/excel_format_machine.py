"""
Ideas on implementation from 8.20 in Python Cookbook 3rd Ed.

Partial implementation of a state machine. It doesn't need to return
to its original state at the moment.
"""
from openpyxl.styles import PatternFill


def flag_difference(first, second):
    pass


key_rules = {
    'Project Name': flag_difference,
    'Joining Date': flag_difference
}


class CellFormatState:
    """
    Base class. When initial class is implemented, it immediately switches
    to a BlankCellFormat class:

        c = CellFormatState()
        c.__class__
        <BlankCellFormat>

    I should use DocTest for this.

    The first step, where the initialiser sets the state for the "next" class
    is where this gets its state machine whiff, but it't not really a state
    machine.
    """

    def __init__(self):
        self.new_state(BlankCellFormat)

    def new_state(self, state):
        self.__class__ = state

    def set_style(self):
        """
        Returns an openpyxl PatternFill object.
        """
        raise NotImplementedError()

    def action(self, *, compare_val=None, this_val=None, key=None):
        """
        Pass data from implementing code into object for comparison and
        processing.
        """
        raise NotImplementedError()

    def compare(self, first, second):
        raise NotImplementedError()

    def clear(self):
        raise NotImplementedError()


class BlankCellFormat(CellFormatState):
    """
    Initial format state of a cell. No cell formatting is implemented.

    Calling action() will change object to an applicable class, depending
    on the parameters to action(), which are passed in from the implementing
    code, and therefore the values that are to be written to the cell or
    compared to.

    If the key parameter matches anything on the key_rules list in this class,
    we know that it's rule should be applied first. After that, if any other
    rule is applied after that, the latter rule gets precedence. (This will
    need to be re-implemented such that you can define any rule you want for
    any key. We still need to decide what overrides what though.)
    """

    def action(self, *, compare_val=None, this_val=None, key=None):
        """
        Takes all data from the data set being evaluated by the Writer class.
        If we have a compare value, we don't do anything unless it is of the
        same type as the value being evaluated by the Writer class. If the
        types do match, then the object is converted into the applicable
        CellFormat class (which are based on CellFormatState.
        """
        if isinstance(this_val, str) and isinstance(compare_val, str):
            self.__class__ = StringCellFormat
            self.compare_val = compare_val
            self.this_val = this_val
            self.key = key
        elif isinstance(this_val, int) and isinstance(compare_val, int):
            self.__class__ = IntegerCellFormat
            self.compare_val = compare_val
            self.this_val = this_val
            self.key = key
        else:
            pass

    def export_rule(self):
        return None


class StringCellFormat(CellFormatState):
    """
    Formatting for a cell that accepts strings is processed here.

    Calling export_rule() by the implementing code requires no parameters
    because you should only be getting here by first calling the base class,
    and calling action() on that with the requisite params. That object is then
    converted to a StringCellFormat object if its compare_val and this_val
    parameters are strings.

    The class variable sets the colour used by the set_style() method.

    This class could be adapted to do a lot more than simply set the background
    in the cell based on the compare_val and this_val data being different.
    """

    rgb = [255, 0, 0]

    def set_style(self, rgb):
        c_value = "{0:02X}{1:02X}{2:02X}".format(*rgb)
        return PatternFill(
            patternType='solid',
            fgColor=c_value,
            bgColor=c_value
        )

    def export_rule(self):
        if self.compare_val != self.this_val:
            return self.set_style(StringCellFormat.rgb)
        else:
            self.__class__ = BlankCellFormat

    def clear(self):
        self.new_state(BlankCellFormat)


class IntegerCellFormat(CellFormatState):
    """
    Formatting for a cell that accepts integers is processed here.

    Calling export_rule() by the implementing code requires no parameters
    because you should only be getting here by first calling the base class,
    and calling action() on that with the requisite params. That object is then
    converted to a IntCellFormat object if its compare_val and this_val
    parameters are integers.

    The class variable sets the colour used by the set_style() method.

    This class could be adapted to do a lot more than simply set the background
    in the cell based on the compare_val and this_val data being different.
    """

    rgb_this_higher = [155, 0, 0]
    rgb_this_lower = [225, 0, 0]

    def set_style(self, rgb):
        c_value = "{0:02X}{1:02X}{2:02X}".format(*rgb)
        return PatternFill(
            patternType='solid',
            fgColor=c_value,
            bgColor=c_value
        )

    def export_rule(self):
        if self.compare_val > self.this_val:
            return self.set_style(IntegerCellFormat.rgb_this_lower)
        elif self.compare_val < self.this_val:
            return self.set_style(IntegerCellFormat.rgb_this_higher)
        else:
            self.__class__ = BlankCellFormat

    def clear(self):
        self.new_state(BlankCellFormat)
