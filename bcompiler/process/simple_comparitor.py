from openpyxl import load_workbook


class BCCell:

    def __init__(self, value, row_num=None, col_num=None, cellref=None):
        self.value = value
        self.row_num = row_num
        self.col_num = col_num
        self.cellref = cellref


class ParsedMaster:

    def __init__(self, master_file):
        self.master_file = master_file
        self._projects = []
        self._project_count = None
        self._key_col = []
        self._wb = load_workbook(self.master_file)
        self._ws = self._wb.active
        self._parse()

    def _parse(self):
        self._projects = [cell.value for cell in self._ws[1][1:]]
        self._projects.sort()
        self._project_count = len(self.projects)
        self._key_col = [cell.value for cell in self._ws['A']]

    @property
    def projects(self):
        return self._projects

    def _create_single_project_tuple(self, column=None, col_index=None):
        if col_index is None:
            col_data = self._ws[column]
            z = list(zip(self._key_col, col_data))
            return [((item[0]), (item[1].value)) for item in z]
        else:
            col_data = []
            for row in self._ws.iter_rows(
                min_row=1,
                max_col=col_index,
                min_col=col_index,
                max_row=len(self._key_col)
            ):
                count = 0
                for cell in row:
                    col_data.append(cell.value)
                    count += 1
            z = list(zip(self._key_col, col_data))
            return [((item[0]), (item[1])) for item in z]

    def _create_dict_all_project_tuples(self):
        pass

    def __repr__(self):
        return "ParsedMaster for {}".format(
            self.master_file
        )

    def get_project_data(self, column=None, col_index=None):
        if col_index:
            data = self._create_single_project_tuple(col_index=col_index)
        else:
            data = self._create_single_project_tuple(column)
        return data


def populate_cells(worksheet, bc_cells=[]):
    """
    Populate a worksheet with bc_cell object data.
    """
    for item in bc_cells:
        if item.cellref:
            worksheet[item.cellref].value = item.value
        else:
            worksheet.cell(
                row=item.row_num, column=item.col_num, value=item.value)
    return worksheet


class SimpleComparitor:
    """
    Simple method of comparing data in two master spreadsheets.
    """

    def __init__(self, masters=[]):
        """
        We want to get a list of master spreadsheets. These are simple
        file-references. The latest master should be master[-1].
        """
        if len(masters) > 2:
            raise ValueError("You can only analyse two spreadsheets.")

        self.masters = masters

    def get_data(self, spreadsheet_file, col):
        wb = load_workbook(spreadsheet_file)
        ws = wb['Constructed BICC Data Master']
        col_a = ws['A']
        col_b = ws[col]
        z = list(zip(col_a, col_b))
        return [((item[0].value), (item[1].value)) for item in z]

    def data(self, index, col):
        return self.get_data(self.masters[index], col)

