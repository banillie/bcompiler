import os

from bcompiler.core import Quarter, Master, Row
from ..utils import logger, ROOT_PATH, CONFIG_FILE, runtime_config

from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl import Workbook
from openpyxl.drawing.line import LineProperties

runtime_config.read(CONFIG_FILE)


def _replace_underscore(name: str):
    return name.replace('/', '_')


def _color_gen():
    for c in [
        'ce5089',
        'ce5650',
        'ce50c8',
        '5050ce',
        '8f50ce',
        '508fce',
        '50ceac',
        '50b1ce',
        '50ce6d'
    ]:
        yield c


def _create_chart(worksheet):
    """Create the fucking chart"""
    chart = ScatterChart()
    chart.varyColors = True
    chart.title = "Financial Analysis"
    chart.style = 1
    chart.height = 10
    chart.width = 20
    chart.x_axis.title = "Node"
    chart.y_axis.title = "Cash"
    chart.legend = None
    chart.x_axis.majorUnit = 0.5
    chart.x_axis.minorGridlines = None
#   chart.y_axis.majorUnit = 200

    xvalues = Reference(worksheet, min_col=1, min_row=3, max_row=4)
    picker = _color_gen()
    for i in range(2, 6):
        values = Reference(worksheet, min_col=i, min_row=2, max_row=4)
        series = Series(values, xvalues, title_from_data=True)
        series.smooth = True
        lineProp = LineProperties(solidFill=next(picker))
        series.graphicalProperties.line = lineProp
        chart.series.append(series)
    worksheet.add_chart(chart, "G1")
    return worksheet


def run(masters_repository_dir, output_path=None):
    wb = Workbook()

    q1 = Quarter(1, 2017)
    q2 = Quarter(2, 2017)

    # TODO - we need a function in here that gleans quarter from the filename
    # of the master

    master_q1 = Master(q1, os.path.join(masters_repository_dir, 'compiled_master_2017-07-18_Q1 Apr - Jun 2017 FOR Q2 COMMISSION DO NOT CHANGE.xlsx'))
    master_q2 = Master(q2, os.path.join(masters_repository_dir, '1718_Q2_master.xlsx'))
    target_keys = [
        'RDEL Total Forecast',
        'CDEL Total Forecast',
        'Non-Gov Total Forecast',
        'Total Forecast SR (20/21)'
    ]

    # projects from latest master
    projects = master_q2.projects


    std = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(std)

    # set up sheets
    for p in projects:
        try:
            ws = wb.create_sheet(_replace_underscore(p))
            start_row = 1
        except AttributeError:
            continue
        else:
            ws.cell(row=start_row, column=1, value=p)
            header = Row(2, start_row + 1, target_keys)
            header.bind(ws)

        for m in [master_q1, master_q2]:
            try:
                p_data = m[p]
            except KeyError:
                logger.warning(f"Cannot find {p}")
                continue
            d = p_data.pull_keys(target_keys, flat=True)
            ws.cell(row=start_row + 2, column=1, value=str(m.quarter))
            r = Row(2, start_row + 2, d)
            r.bind(ws)

            start_row += 1

        ws = _create_chart(ws)

    if output_path:
        wb.save(os.path.join(output_path[0], 'financial_analysis.xlsx'))
        logger.info(f"Saved financial_analysis.xlsx to {output_path}")
    else:
        output_path = os.path.join(ROOT_PATH, 'output')
        wb.save(os.path.join(output_path, 'financial_analysis.xlsx'))
        logger.info(f"Saved financial_analysis.xlsx to {output_path}")


if __name__ == '__main__':
    run('/tmp/master_repo')
